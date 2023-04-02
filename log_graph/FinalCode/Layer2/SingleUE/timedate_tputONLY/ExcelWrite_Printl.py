import openpyxl, string
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def writeexcel(result):
    f = open(result, 'r+')  # open text
    #########if load excel file ########################
    excel = openpyxl.Workbook()
    sheet = excel.worksheets
    line = f.readline();  # read text
    sheet2=excel.create_sheet("Mysheet1", 1) 
    new=[]
    count =0
    while line:
        list123 = line.split()  # convert
        if "=" in line:            
            pass  
        elif "-" not in line and "#" not in line:   
        #else:    
            if count <=2:
                if list123[1] == 'Tput':
                    sheet[0].append(list123)  # write into excel
                else:
                    list123[1] = float(list123[1])
                    sheet[0].append(list123)  # write into excel         
        line = f.readline()  # read next line

    #excel.save(excelfilename+'.xlsx')
    #excel.save('result.xlsx')
    excel.save(excelfilename+".xlsx")
resultfilename=input("Please enter your txt file:")
excelfilename=input("Please enter your xlsx file:")

#resultfilename="result.txt"
writeexcel(resultfilename)




