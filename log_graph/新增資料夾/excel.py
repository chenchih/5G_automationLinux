import openpyxl, string
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def excelconvertMAC(result):
    f = open(result, 'r+')  # open text
    #########if load excel file ########################
    # excel=openpyxl.load_workbook(r'D:\\test\\test.xlsx') #open excel
    excel = openpyxl.Workbook()
    sheet = excel.worksheets
    line = f.readline();  # read text
    while line:
        #list123 = []
        #list123 = line.split(sep=' ')  # convert,
        list123 = line.split()  # convert
        #remove below 
        #for i in range(0, len(list123)):  # remove space
        #    list123[i] = list123[i].strip('\n') 
        #sheet[0].append(list123)  # write into excel

        if list123[1] == 'Tput':
            sheet[0].append(list123)  # write into excel
        elif list123[2] == 'RbNum':
            sheet[0].append(list123)  # write into excel
        elif list123[3] == 'MCS':
            sheet[0].append(list123)  # write into excel
        elif list123[4] == 'PdschBler':
            sheet[0].append(list123)  # write into excel
        elif list123[5] == 'nonWPdschBler':
            sheet[0].append(list123)  # write into excel
        else:
            list123[1] = float(list123[1])
            list123[2] = float(list123[2])
            list123[3] = float(list123[3])
            list123[4] = float(list123[4])
            list123[5] = float(list123[5])
            sheet[0].append(list123)  # write into excel
    
#adjust the column width 
        column = 1
        while column < 6:
            i = get_column_letter(column)  
            #print(i)         
            sheet[0].column_dimensions[i].width = 25    
              
            column += 1
        
        sheet[0]['A1'].font = Font(size = 14, bold = True)
        sheet[0]['B1'].font = Font(size = 14, bold = True)
        sheet[0]['C1'].font = Font(size = 14, bold = True)
        sheet[0]['D1'].font = Font(size = 14, bold = True)
        sheet[0]['E1'].font = Font(size = 14, bold = True)
        sheet[0]['F1'].font = Font(size = 14, bold = True)
        #debug use
        #sheet1 = excel.worksheets[0]        
        #sheet1['A1'] .font = Font(size = 24, bold = True)
        #sheet1['A1'] = 'Hello Python, Hello Excel.'
        line = f.readline()  # read next line
    #excel.save(r'C:\MAC_add_Submission\MAC_List.xlsx')
    excel.save('result.xlsx')
resultfilename=input("please enter your report txt file name: ")
#resultfilename="result.txt"
excelconvertMAC(resultfilename)




"""
    while line:
        list123 = line.split()  # create a list
        print(type(list123))
        if list123[1] == 'Tput':
            sheet[0].append(list123)  # write into excel
        elif list123[2] == 'Tput2':
            sheet[0].append(list123)  # write into excel
        else:
            list123[1] = float(list123[1])
            list123[2] = float(list123[1])
            sheet[0].append(list123)  # write into excel   
"""  