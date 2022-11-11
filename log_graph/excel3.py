import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
 
def convert_cell(text):
    """Try to convert cell text to object of appropriate type"""
    # Currently the only type we convert are floats
    try:
        return float(text)
    except:
        pass
    return text
 
def text_file_to_excel(src, dest):
    """Read spreadsheet like thing from string file and convet to excel"""
    wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]
 
    with open(src, "r") as src:
        # First line is column headings
        line = next(src).split()
        print(src)
        sheet.append(line)
        for column in range(1, len(line)+1):
            sheet.cell(1, column).font = Font(size=14, bold=True)
            sheet.column_dimensions[get_column_letter(column)].width = 25
 
        # Copy remaining rows converting number strings to numbers.
        for line in src:
            print(line)
            sheet.append([convert_cell(text) for text in line.split()])
    wb.save(dest)
  
text_file_to_excel("result.txt", "test.xlsx")