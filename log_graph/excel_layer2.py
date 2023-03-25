import openpyxl, string
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def excelconvertMAC(result):
    f = open(result, 'r+')  # open text
    #########if load excel file ########################
    # excel=openpyxl.load_workbook(r'D:\\test\\test.xlsx') #open excel
    excel = openpyxl.Workbook()
    sheet = excel.worksheets
    line = f.readline();  # read text

    while line:

        list123 = line.split()  # convert
        
        if "=" in line:
            pass
            
            #list123 = []
            #list123 = line.split(sep=' ')  # convert,
        else:
        #print(line)
        #if not "=" in line:


            if list123[1] == 'Tput':
            
                sheet[0].append(list123)  # write into excel
                
                
                
                
            elif list123[2] == 'RbNum':
                sheet[0].append(list123)  # write into excel
            elif list123[3] == 'MCS':
                sheet[0].append(list123)  # write into excel
            elif list123[4] == 'PdschBler':
                sheet[0].append(list123)  # write into excel
            elif list123[5] == 'nonWPdschBler':
                sheet[0].append(list123)  # write into excel
            
            else:
                
                list123[1] = float(list123[1])
                list123[2] = float(list123[2])
                list123[3] = float(list123[3])
                list123[4] = float(list123[4])
                list123[5] = float(list123[5])
                sheet[0].append(list123)  # write into excel
               
                #excel cell's font
                sheet[0]['A1'] .font = Font(size = 24, bold = True)
                sheet[0]['B1'].font = Font(size = 14, bold = True)
                sheet[0]['C1'].font = Font(size = 14, bold = True)
                sheet[0]['D1'].font = Font(size = 14, bold = True)
                sheet[0]['E1'].font = Font(size = 14, bold = True)
                sheet[0]['F1'].font = Font(size = 14, bold = True)
                
        #adjust the column width 
        column = 1
        while column < 6:
            i = get_column_letter(column)  
            #print(i)         
            sheet[0].column_dimensions[i].width = 25    
              
            column += 1
        line = f.readline()  # read next line
    excel.save(excelfilename+'.xlsx')
resultfilename=input("please enter your report txt file name: ")
excelfilename=input("please enter save excel file name: ")
excelconvertMAC(resultfilename)




