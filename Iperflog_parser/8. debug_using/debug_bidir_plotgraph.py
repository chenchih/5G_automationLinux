import re
from datetime import datetime
import time
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

def parse_log_file_rx_tx(input_file):


    with open(input_file, 'r') as infile:
        for line in infile:
            match = re.match(r"(\w{3} \w{3}\s+\d{1,2} \d{2}:\d{2}:\d{2} \d{4}) \[SUM]\[(RX-C|TX-C)\].*?([\d\.]+) (bits|Mbits|Gbits)/sec", line)
            if match:
                date_str = match.group(1)#datetime
                rx_tx = match.group(2)#TX-C and RX-C
                transfer_str = match.group(3)#tput value
                unit = match.group(4)#unit 
                
                # Increment the appropriate counter
                #print(date_str, rx_tx, transfer_str, unit)
                
                date_obj = datetime.strptime(date_str, "%a %b %d %H:%M:%S %Y")
                formatted_date = date_obj.strftime("%Y%m%d_%H:%M:%S")
                transfer_value = float(transfer_str)                                  
                if unit == "Mbits":
                    transfer_value /= 1000.0
                elif unit == "bits":
                    transfer_value /= 1000000000.0    
                
                
                if rx_tx == "RX-C":
                    rx_data.append([formatted_date, transfer_value, "gbps"])
                else: # TX-C
                    tx_data.append([formatted_date, transfer_value, "gbps"])
                #print(formatted_date, transfer_value, unit)
    print(f"Parsing complete. Found {len(rx_data)} RX entries and {len(tx_data)} TX entries.")
    return rx_data, tx_data    

def write_data_to_excel_rx_tx(sheet, data):
    """Writes data rows to the specified Excel sheet."""
    header = ["Datetime", "Tput", "Unit"]
    sheet.append(header)
    if data: # Only show progress bar if there's data
        for row in data:
            sheet.append(row)
    else:
        sheet.append(["No data found for this category."])

def adjust_column_width(sheet):
    """Adjusts the width of columns based on content (simple version for col A)."""
    datetime_column_letter = 'A' # Assuming Datetime is always column A
    max_length = 0
    # Iterate through cells in the column, skipping the header if desired
    for cell in sheet[datetime_column_letter]:
         try:
             # Check if cell value is not None before converting to string
             if cell.value is not None:
                 cell_len = len(str(cell.value))
                 if cell_len > max_length:
                     max_length = cell_len
         except TypeError:
             pass # Ignore type errors if cell value is not string-convertible
    # Set a reasonable width (max length + buffer)
    adjusted_width = max_length + 2
    sheet.column_dimensions[datetime_column_letter].width = adjusted_width
    # You could add similar logic for other columns (B, C) if needed
    sheet.column_dimensions['B'].width = 15 # Example fixed width for Tput
    sheet.column_dimensions['C'].width = 10 # Example fixed width for Unit

def plot_single_throughput(data, direction_label, output_image_file):
    times = [datetime.strptime(row[0], "%Y%m%d_%H:%M:%S") for row in data]
    tput = [row[1] for row in data]
    fig, ax = plt.subplots(figsize=(18, 8)) # Use subplots for more control, slightly wider figsize

    # Plot data
    ax.plot(times, tput, label=f'{direction_label} Throughput (gbps)', marker='.', linestyle='-', markersize=5)
    # --- Formatting ---
    # Set Title and Labels with increased font size
    ax.set_title(f'{direction_label} Throughput', fontsize=16)
    ax.set_xlabel("Time", fontsize=14, labelpad=25)
    ax.set_ylabel("Throughput (gbps)", fontsize=14)
    # Grid
    ax.grid(True, which='major', linestyle='--', linewidth=0.7)
    ax.grid(True, which='minor', linestyle=':', linewidth=0.4) # Fainter minor grid
    # Legend
    ax.legend(fontsize=12)        
    # --- Y-axis Limits ---
    ax.set_ylim(bottom=0)  # Set the lower limit of the y-axis to 0
    # --- X-axis Time Formatting ---
    # Set major locator and formatter for dates
    ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=5, maxticks=50)) # Adjust tick density
    #ax.xaxis.set_major_locator(mdates.MinuteLocator(interval=30)) # Minute Intervals, 
    #ax.xaxis.set_major_locator(mdates.HourLocator(interval=1)) # Hourly Intervals, show a tick every hour
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y%m%d_%H:%M:%S')) # Multi-line format
    
    # Increase tick label size and rotate
    plt.setp(ax.get_xticklabels(), rotation=90, ha='right', fontsize=12)
    plt.setp(ax.get_yticklabels(), fontsize=12)
    
    ax.tick_params(axis='x', pad=10) 

    # Adjust layout to prevent labels overlapping
    fig.tight_layout()
    #plt.show()
    # --- Save Output ---
    print(f"Saving plot to {output_image_file}...")
    plt.savefig(output_image_file, dpi=300) # Save the plot (dpi for resolution)
    plt.close(fig) # Close the plot figure to free memory
    print(f"{direction_label} plot saved successfully.")
    return True # Indicate plotting was successful
    
    
def process_log_file_to_excel_rx_tx(rx_data, tx_data, output_excel_file):
    workbook = openpyxl.Workbook()
    if rx_data:
        rx_sheet = workbook.create_sheet("RX-DL")
        write_data_to_excel_rx_tx(rx_sheet, rx_data)
        adjust_column_width(rx_sheet)
    if tx_data:
        tx_sheet = workbook.create_sheet("TX-UL")
        write_data_to_excel_rx_tx(tx_sheet, tx_data)
        adjust_column_width(tx_sheet)
# Remove the default 'Sheet' if it exists and we added others
    if ('Sheet' in workbook.sheetnames and (rx_data or tx_data)):
        del workbook['Sheet']

    print('--------------------------------------------------')
    print(f"Saving Excel file to {output_excel_file}... Please wait.")
    workbook.save(output_excel_file)
    print(f"Data successfully written to {output_excel_file}")

    
###########################################
rx_data = []
tx_data = []
input_file= input("Please enter your elog FileName: ")
output_excel_file_path = 'output.xlsx'
output_rx_plot_filename = "output_RX-DL_plot.png"
output_tx_plot_filename = "output_TX-UL_plot.png"
try:
    #process_log_file_to_excel_combined(input_file, output_excel_file_path)
    #parse_log_file(input_file)
    rx_data, tx_data = parse_log_file_rx_tx(input_file)
    #if rx_data or tx_data:
    process_log_file_to_excel_rx_tx(rx_data, tx_data, output_excel_file_path)
    if rx_data:
        plot_single_throughput(rx_data, "RX-DL", output_rx_plot_filename)
    if tx_data:
        plot_single_throughput(tx_data, "TX-DL", output_tx_plot_filename)
        
except SystemExit:
    print('\nError: Your log file contains Gbits/sec. This script only captures Mbits/sec log files.')
