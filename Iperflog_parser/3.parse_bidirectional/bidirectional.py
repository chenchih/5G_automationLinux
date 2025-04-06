import re, os, shutil
from datetime import datetime
import openpyxl
from tqdm import tqdm

def process_log_file_to_excel_rx_tx(input_file, output_excel_file):
    try:
        rx_data, tx_data = parse_log_file_rx_tx(input_file)
        if rx_data or tx_data:
            workbook = openpyxl.Workbook()
            if rx_data:
                rx_sheet = workbook.create_sheet("RX-DL")
                write_data_to_excel_rx_tx(rx_sheet, rx_data)
                adjust_column_width(rx_sheet)
            if tx_data:
                tx_sheet = workbook.create_sheet("TX-UL")
                write_data_to_excel_rx_tx(tx_sheet, tx_data)
                adjust_column_width(tx_sheet)
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            print('--------------------------------------------------')
            print("Saving Excel file... Please wait.") #added print 
            workbook.save(output_excel_file)
            print(f"Data written to {output_excel_file}")
            #Remove calculate_and_print_duration from here
            #if rx_data:
                #calculate_and_print_duration(rx_data, 'RX-DL')
            #if tx_data:
             	#calculate_and_print_duration(tx_data,'TX-UL')
        else:
            print("No RX-C or TX-C lines found in the input file.")
    except FileNotFoundError:
        print(f"Error: Input file '{input_file}' not found.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        
def parse_log_file_rx_tx(input_file):
    rx_data = []
    tx_data = []
    with open(input_file, 'rb') as infile: # Open in binary mode
        total_size = os.path.getsize(input_file)
        bytes_read = 0

        with tqdm(total=total_size, unit='B', unit_scale=True, desc="Reading File") as pbar:
            for line in infile:
                bytes_read += len(line)
                pbar.update(len(line))
                try:
                    line_str = line.decode('utf-8', errors='ignore')
                    #match = re.match(r"(\w{3} \w{3} \d{2} \d{2}:\d{2}:\d{2} \d{4}) \[SUM]\[(RX-C|TX-C)\].*?([\d\.]+) (M|G)bits/sec", line_str)
                    #match = re.match(r"(\w{3} \w{3} \d{2} \d{2}:\d{2}:\d{2} \d{4}) \[SUM]\[(RX-C|TX-C)\].*?([\d\.]+) (bits|Mbits|Gbits)/sec", line_str)           
                    match = re.match(r"(\w{3} \w{3}\s+\d{1,2} \d{2}:\d{2}:\d{2} \d{4}) \[SUM]\[(RX-C|TX-C)\].*?([\d\.]+) (bits|Mbits|Gbits)/sec", line_str)
                    if match:
                        date_str = match.group(1)#datetime
                        rx_tx = match.group(2)#TX-C and RX-C
                        transfer_str = match.group(3)#tput value
                        unit = match.group(4)#unit 
                        try:
                            date_obj = datetime.strptime(date_str, "%a %b %d %H:%M:%S %Y")
                            formatted_date = date_obj.strftime("%Y%m%d_%H:%M:%S")
                            transfer_value = float(transfer_str)
                            #if unit == "M":
                                #transfer_value /= 1000.0
                            # Convert to Gbits/sec
                            if unit == "Mbits":
                                transfer_value /= 1000.0
                            elif unit == "bits":
                                transfer_value /= 1000000000.0
                            # Append data    
                            if rx_tx == "RX-C":
                                rx_data.append([formatted_date, transfer_value, "gbps"])
                            else:
                                tx_data.append([formatted_date, transfer_value, "gbps"])
                        except ValueError:
                            print(f"Warning: Invalid data format in line: '{line_str}'. Skipping.")
                except UnicodeDecodeError:
                    print(f"Warning: UnicodeDecodeError in Line: {line}")
    return rx_data, tx_data


#Writes data to the Excel sheet.
def write_data_to_excel_rx_tx(sheet, data):
    
    header = ["Datetime", "Tput", "Unit"]
    sheet.append(header)
    with tqdm(total=len(data), desc=f"Writing to {sheet.title}") as pbar:
        for row in data:
            sheet.append(row)
            pbar.update(1)
            
#Adjusts the first column's width.
def adjust_column_width(sheet):

    datetime_column = sheet['A']
    max_length = 0
    for cell in datetime_column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except TypeError:
            pass
    adjusted_width = max_length + 2
    sheet.column_dimensions['A'].width = adjusted_width

def moving_file(logfile, excelfile):
    now = datetime.now()
    folder_name = now.strftime("%Y-%m-%d_%H-%M-%S")  # Format: YYYY-MM-DD_HH-MM-SS

    os.makedirs(folder_name, exist_ok=True)  # Create folder, no error if it exists
    print(f"Folder '{folder_name}' created.")

    # Move the file into the folder
    destination_path = os.path.join(folder_name, excelfile)
    shutil.move(excelfile, destination_path)
    print(f"File '{excelfile}' moved to '{destination_path}'.")

def calculate_and_print_duration(data, item):
    """Calculates and prints the running duration."""
    if not data:
        return

    start_time_str = data[0][0]
    end_time_str = data[-1][0]

    start_time = datetime.strptime(start_time_str, "%Y%m%d_%H:%M:%S")
    end_time = datetime.strptime(end_time_str, "%Y%m%d_%H:%M:%S")

    duration = end_time - start_time
    total_seconds = int(duration.total_seconds())

    days = total_seconds // (24 * 3600)
    remaining_seconds = total_seconds % (24 * 3600)
    hours = remaining_seconds // 3600
    remaining_seconds %= 3600
    minutes = remaining_seconds // 60

    print(f"{item} => Running duration: {days} days, {hours} hours, {minutes} minutes")
    print(f"{item} => Converted hours: {days * 24 + hours} hours")
    


input_file_path = input('Enter your filename: ')
output_excel_file_path = (input('save exel file name (enter for default) : ') )
print('--------------------------------------------------')
if output_excel_file_path == '':
    output_excel_file_path = "output_rx_tx.xlsx"
else:
    output_excel_file_path=output_excel_file_path+'.xlsx'
    
rx_data, tx_data = parse_log_file_rx_tx(input_file_path) # parse the data first    
process_log_file_to_excel_rx_tx(input_file_path, output_excel_file_path)
moving_file(input_file_path, output_excel_file_path)
print('--------------------------------------------------')
if rx_data:
    calculate_and_print_duration(rx_data, 'RX-DL')
if tx_data:
    calculate_and_print_duration(tx_data, 'TX-UL')