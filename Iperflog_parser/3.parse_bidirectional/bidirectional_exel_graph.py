import re
import os
import shutil
from datetime import datetime
import openpyxl
from tqdm import tqdm
import matplotlib.pyplot as plt
import matplotlib.dates as mdates # Needed for date formatting and locators

def parse_log_file_rx_tx(input_file):
    """Parses the log file to extract RX and TX throughput data."""
    rx_data = []
    tx_data = []
    print(f"Attempting to open log file: {input_file}")
    try:
        with open(input_file, 'rb') as infile: # Open in binary mode
            total_size = os.path.getsize(input_file)
            bytes_read = 0
            print("Parsing log file...")
            with tqdm(total=total_size, unit='B', unit_scale=True, desc="Reading Log") as pbar:
                for line in infile:
                    bytes_read += len(line)
                    pbar.update(len(line))
                    try:
                        line_str = line.decode('utf-8', errors='ignore')
                        # Regex to capture date, RX/TX, value, and unit
                        match = re.match(r"(\w{3} \w{3}\s+\d{1,2} \d{2}:\d{2}:\d{2} \d{4}) \[SUM]\[(RX-C|TX-C)\].*?([\d\.]+) (bits|Mbits|Gbits)/sec", line_str)
                        if match:
                            date_str = match.group(1) # datetime string
                            rx_tx = match.group(2)    # TX-C or RX-C
                            transfer_str = match.group(3) # throughput value string
                            unit = match.group(4)     # unit (bits, Mbits, Gbits)
                            try:
                                date_obj = datetime.strptime(date_str, "%a %b %d %H:%M:%S %Y")
                                formatted_date = date_obj.strftime("%Y%m%d_%H:%M:%S")
                                transfer_value = float(transfer_str)

                                # Convert all to Gbits/sec for consistency
                                if unit == "Mbits":
                                    transfer_value /= 1000.0
                                elif unit == "bits":
                                    transfer_value /= 1000000000.0
                                # elif unit == "Gbits": # No conversion needed
                                #     pass

                                # Append data
                                if rx_tx == "RX-C":
                                    rx_data.append([formatted_date, transfer_value, "gbps"])
                                else: # TX-C
                                    tx_data.append([formatted_date, transfer_value, "gbps"])
                            except ValueError:
                                print(f"Warning: Invalid data format in line (ValueError): '{line_str.strip()}'. Skipping.")
                            except Exception as e_inner:
                                print(f"Warning: Error processing matched line: '{line_str.strip()}' - {e_inner}. Skipping.")

                    except UnicodeDecodeError:
                        print(f"Warning: UnicodeDecodeError reading a line. Skipping binary data.")
                    except Exception as e_outer:
                         print(f"Warning: Unexpected error reading line: {e_outer}")


    except FileNotFoundError:
        print(f"Error: Input file '{input_file}' not found.")
        return [], [] # Return empty lists if file not found
    except Exception as e:
        print(f"An unexpected error occurred during parsing: {e}")
        return [], [] # Return empty lists on other errors

    print(f"Parsing complete. Found {len(rx_data)} RX entries and {len(tx_data)} TX entries.")
    return rx_data, tx_data

def write_data_to_excel_rx_tx(sheet, data):
    """Writes data rows to the specified Excel sheet."""
    header = ["Datetime", "Tput", "Unit"]
    sheet.append(header)
    if data: # Only show progress bar if there's data
      with tqdm(total=len(data), desc=f"Writing {sheet.title}", unit=" rows") as pbar:
          for row in data:
              sheet.append(row)
              pbar.update(1)
    else:
        sheet.append(["No data found for this category."])

def adjust_column_width(sheet):
    """Adjusts the width of columns based on content (simple version for col A)."""
    datetime_column_letter = 'A' # Assuming Datetime is always column A
    max_length = 0
    # Iterate through cells in the column, skipping the header if desired
    for cell in sheet[datetime_column_letter]:
         try:
             # Check if cell value is not None before converting to string
             if cell.value is not None:
                 cell_len = len(str(cell.value))
                 if cell_len > max_length:
                     max_length = cell_len
         except TypeError:
             pass # Ignore type errors if cell value is not string-convertible
    # Set a reasonable width (max length + buffer)
    adjusted_width = max_length + 2
    sheet.column_dimensions[datetime_column_letter].width = adjusted_width
    # You could add similar logic for other columns (B, C) if needed
    sheet.column_dimensions['B'].width = 15 # Example fixed width for Tput
    sheet.column_dimensions['C'].width = 10 # Example fixed width for Unit

def process_log_file_to_excel_rx_tx(rx_data, tx_data, output_excel_file):
    """Creates an Excel file and writes RX and TX data to separate sheets."""
    try:
        if not rx_data and not tx_data:
             print("No RX-C or TX-C data found to write to Excel.")
             return # Exit function if no data

        workbook = openpyxl.Workbook()

        if rx_data:
            rx_sheet = workbook.create_sheet("RX-DL")
            write_data_to_excel_rx_tx(rx_sheet, rx_data)
            adjust_column_width(rx_sheet)

        if tx_data:
            tx_sheet = workbook.create_sheet("TX-UL")
            write_data_to_excel_rx_tx(tx_sheet, tx_data)
            adjust_column_width(tx_sheet)

        # Remove the default 'Sheet' if it exists and we added others
        if ('Sheet' in workbook.sheetnames and (rx_data or tx_data)):
            del workbook['Sheet']

        print('--------------------------------------------------')
        print(f"Saving Excel file to {output_excel_file}... Please wait.")
        workbook.save(output_excel_file)
        print(f"Data successfully written to {output_excel_file}")

    except Exception as e:
        print(f"An unexpected error occurred during Excel processing: {e}")

def calculate_and_print_duration(data, item):
    """Calculates and prints the running duration from the first to last timestamp."""
    if not data or len(data) < 2: # Need at least two points for duration
        print(f"{item} => Not enough data points to calculate duration.")
        return

    try:
        start_time_str = data[0][0]
        end_time_str = data[-1][0]

        start_time = datetime.strptime(start_time_str, "%Y%m%d_%H:%M:%S")
        end_time = datetime.strptime(end_time_str, "%Y%m%d_%H:%M:%S")

        duration = end_time - start_time
        total_seconds = int(duration.total_seconds())

        if total_seconds < 0:
             print(f"{item} => Warning: End time is earlier than start time. Cannot calculate duration.")
             return

        days = total_seconds // (24 * 3600)
        remaining_seconds = total_seconds % (24 * 3600)
        hours = remaining_seconds // 3600
        remaining_seconds %= 3600
        minutes = remaining_seconds // 60
        seconds = remaining_seconds % 60 # Added seconds for more detail

        print(f"{item} => Start Time: {start_time}")
        print(f"{item} => End Time:   {end_time}")
        print(f"{item} => Running duration: {days} days, {hours} hours, {minutes} minutes, {seconds} seconds (Total: {total_seconds}s)")

    except ValueError:
        print(f"{item} => Error parsing date/time for duration calculation.")
    except Exception as e:
        print(f"{item} => An unexpected error occurred during duration calculation: {e}")




def moving_files(timestamp_folder, files_to_move):
    """Moves a list of files into the specified timestamped folder."""
    try:
        os.makedirs(timestamp_folder, exist_ok=True)
        print(f"Ensured folder '{timestamp_folder}' exists.")

        moved_count = 0
        for file_path in files_to_move:
            if file_path and os.path.exists(file_path):
                try:
                    base_filename = os.path.basename(file_path)
                    destination_path = os.path.join(timestamp_folder, base_filename)
                    shutil.move(file_path, destination_path)
                    print(f"File '{base_filename}' moved to '{destination_path}'.")
                    moved_count += 1
                except Exception as e_move:
                    print(f"Error moving file '{file_path}': {e_move}")
            elif file_path:
                 print(f"File '{file_path}' not found, skipping move.")

        if moved_count == 0:
            print("No files were found to move.")

    except Exception as e_folder:
        print(f"Error creating or accessing folder '{timestamp_folder}': {e_folder}")

# --- UPDATED Plotting Function ---

        
def plot_single_throughput(data, direction_label, output_image_file):
    """
    Generates and saves a throughput plot for a single direction (RX or TX).
    Increases font sizes for better readability.
    """
    print('--------------------------------------------------')
    print(f"Generating plot for {direction_label}...")
    if not data:
        print(f"No {direction_label} data available to plot.")
        return False # Indicate plotting was not done

    try:
        # Extract times and throughput, converting times to datetime objects
        times = [datetime.strptime(row[0], "%Y%m%d_%H:%M:%S") for row in data]
        tput = [row[1] for row in data]

        if not times: # Check if data extraction was successful
             print(f"No valid time data points found for {direction_label}.")
             return False

        # --- Create Plot ---
        fig, ax = plt.subplots(figsize=(18, 8)) # Use subplots for more control, slightly wider figsize

        # Plot data
        ax.plot(times, tput, label=f'{direction_label} Throughput (gbps)', marker='.', linestyle='-', markersize=5)

        # --- Formatting ---
        # Set Title and Labels with increased font size
        ax.set_title(f'{direction_label} Throughput', fontsize=16)
        ax.set_xlabel("Time", fontsize=14)
        ax.set_ylabel("Throughput (gbps)", fontsize=14)

        # Grid
        ax.grid(True, which='major', linestyle='--', linewidth=0.7)
        ax.grid(True, which='minor', linestyle=':', linewidth=0.4) # Fainter minor grid

        # Legend
        ax.legend(fontsize=12)
        
        # --- Y-axis Limits ---
        ax.set_ylim(bottom=0)  # Set the lower limit of the y-axis to 0


        # --- X-axis Time Formatting ---
        # Set major locator and formatter for dates
        ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=5, maxticks=50)) # Adjust tick density
        #ax.xaxis.set_major_locator(mdates.MinuteLocator(interval=30)) # Minute Intervals, 
        #ax.xaxis.set_major_locator(mdates.HourLocator(interval=1)) # Hourly Intervals, show a tick every hour
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y%m%d_%H:%M:%S')) # Multi-line format
        
        # Increase tick label size and rotate
        plt.setp(ax.get_xticklabels(), rotation=90, ha='right', fontsize=12)
        plt.setp(ax.get_yticklabels(), fontsize=12)
        

        # Adjust layout to prevent labels overlapping
        fig.tight_layout()

        # --- Save Output ---
        print(f"Saving plot to {output_image_file}...")
        plt.savefig(output_image_file, dpi=300) # Save the plot (dpi for resolution)
        plt.close(fig) # Close the plot figure to free memory
        print(f"{direction_label} plot saved successfully.")
        return True # Indicate plotting was successful

    except ValueError as ve:
         print(f"Error processing/plotting {direction_label} data (ValueError): {ve}")
         plt.close() # Ensure figure is closed on error
         return False
    except Exception as e:
        print(f"An unexpected error occurred during {direction_label} plotting: {e}")
        plt.close() # Ensure figure is closed on error
        return False


# --- Main Execution Block ---

if __name__ == "__main__":
    datename = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    input_file_path = input('Enter your log filename (e.g., iperf3.log): ')
    output_base_name = input('Enter base name for output files (press Enter for default): ')

    print('--------------------------------------------------')

    # Define output filenames
    if not output_base_name:
        # Generate default base name using current time and potentially part of input filename
        log_basename = os.path.splitext(os.path.basename(input_file_path))[0] #gett the filename only
        output_base_name = f"{datename}_{log_basename}_analysis"
        

    output_excel_filename = output_base_name + ".xlsx"
    # Define separate plot filenames
    output_rx_plot_filename = output_base_name + "_RX-DL_plot.png"
    output_tx_plot_filename = output_base_name + "_TX-UL_plot.png"

    #output_folder_name = output_base_name + "_Results" # Folder to move files into
    output_folder_name =  f"{datename.split('_')[0]}_{log_basename}_analysis"
    # 1. Parse the log file
    rx_data, tx_data = parse_log_file_rx_tx(input_file_path)

    files_generated = [] # Keep track of files we successfully create

    # 2. Write data to Excel
    if rx_data or tx_data:
        process_log_file_to_excel_rx_tx(rx_data, tx_data, output_excel_filename)
        if os.path.exists(output_excel_filename): # Verify file exists before adding
             files_generated.append(output_excel_filename)
        else:
             print(f"Warning: Excel file {output_excel_filename} was expected but not found.")
    else:
         print("Skipping Excel generation as no data was parsed.")

    # 3. Plot RX data (if available)    
    if rx_data:
        
        rx_plot_successful = plot_single_throughput(rx_data, "RX-DL", output_rx_plot_filename)
        if rx_plot_successful:
             files_generated.append(output_rx_plot_filename)
        else:
            print("Skipping RX plot moving as it was not generated successfully.")
    else:
        print("No RX data to plot.")

    # 4. Plot TX data (if available)
    if tx_data:
        tx_plot_successful = plot_single_throughput(tx_data, "TX-UL", output_tx_plot_filename)
        if tx_plot_successful:
             files_generated.append(output_tx_plot_filename)
        else:
            print("Skipping TX plot moving as it was not generated successfully.")
    else:
        print("No TX data to plot.")
        


    # 5. Move generated files to the results folder
    print('--------------------------------------------------')
    print(f"Moving generated files to folder: {output_folder_name}")
    files_to_move = [f for f in files_generated if f] # Filter out None/empty
    if files_to_move:
        moving_files(output_folder_name, files_to_move)
    else:
        print("No files were generated or found to move.")


    # 6. Calculate and print durations
    print('--------------------------------------------------')
    print("Calculating Durations:")
    # Reuse the same calculate_and_print_duration function
    if rx_data:
        calculate_and_print_duration(rx_data, 'RX-DL')
    if tx_data:
        calculate_and_print_duration(tx_data, 'TX-UL')
    if not rx_data and not tx_data:
        print("No data found to calculate durations.")

    print('--------------------------------------------------')
    print("Script finished.")
