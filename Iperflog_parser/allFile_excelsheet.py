import re
from datetime import datetime
import openpyxl
import os
from openpyxl.styles import Alignment
from tqdm import tqdm

def process_log_files_to_excel(directory_path, output_excel_file):
    """
    Processes multiple log files in a directory, extracts SUM lines, converts transfer rates to Gbits/sec,
    formats data, and writes to separate sheets in an Excel file.

    Args:
        directory_path: Path to the directory containing the log files.
        output_excel_file: Path to the output Excel file.
    """
    try:
        workbook = openpyxl.Workbook()
        
        file_list = [f for f in os.listdir(directory_path) if f.endswith((".txt", ".log"))]  # Get .txt and .log files
        # Delete the default sheet if it exists
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        if not file_list:
            print(f"No .txt or .log files found in directory: {directory_path}")
            return

        with tqdm(total=len(file_list), desc="Processing Files") as pbar_files:
            for filename in file_list:
                input_file_path = os.path.join(directory_path, filename)
                sheet_name = os.path.splitext(filename)[0]  # Use filename (without extension) for sheet name

                # Create a new sheet
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]  # If sheet exists, use it
                else:
                    sheet = workbook.create_sheet(title=sheet_name)  # Create new sheet

                sheet.append(["Datetime", "Tput", "Gbits/sec"])  # Header row

                with open(input_file_path, 'r') as infile:
                    for line in infile:
                        if "[SUM]" in line:
                            # Modified regex to capture both Mbits/sec and Gbits/sec
                            match = re.match(
                                r"(\w{3} \w{3} \d{2} \d{2}:\d{2}:\d{2} \d{4}) \[SUM\].*?([\d\.]+) (M|G)bits/sec",
                                line)
                            if match:
                                date_str = match.group(1)
                                transfer_str = match.group(2)
                                unit = match.group(3)

                                try:
                                    date_obj = datetime.strptime(date_str, "%a %b %d %H:%M:%S %Y")
                                    formatted_date = date_obj.strftime("%Y%m%d_%H:%M:%S")

                                    # Convert to Gbits/sec
                                    transfer_value = float(transfer_str)
                                    if unit == "M":
                                        transfer_value /= 1000.0  # Convert Mbits to Gbits

                                    sheet.append([formatted_date, transfer_value])
                                except ValueError:
                                    print(f"Warning: Invalid data format in line: '{line}'. Skipping.")

                # Adjust column width for 'Datetime' (in each sheet)
                datetime_column = sheet['A']
                max_length = 0
                for cell in datetime_column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except TypeError:
                        pass

                adjusted_width = max_length + 2
                sheet.column_dimensions['A'].width = adjusted_width
                pbar_files.update(1)  # Update progress bar for files

        workbook.save(output_excel_file)
        print(f"Data written to {output_excel_file}")

    except FileNotFoundError:
        print(f"Error: Directory '{directory_path}' not found.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

# Example usage:
directory_path = "."  # Current directory (you can change this)
output_excel_file_path = "output.xlsx"

process_log_files_to_excel(directory_path, output_excel_file_path)