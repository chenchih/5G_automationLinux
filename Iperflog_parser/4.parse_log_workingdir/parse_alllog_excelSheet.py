import re
import sys
from datetime import datetime
import openpyxl
import os
from openpyxl.styles import Alignment
from tqdm import tqdm
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np

def process_log_files_to_excel(directory_path, output_excel_file):

    try:
        workbook = openpyxl.Workbook()

        file_list = [f for f in os.listdir(directory_path) if f.endswith((".txt", ".log"))]  # Get .txt and .log files
        # Delete the default sheet if it exists
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        if not file_list:
            print(f"No .txt or .log files found in directory: {directory_path}")
            return

        with tqdm(total=len(file_list), desc="Processing Files") as pbar_files:
            for filename in file_list:
                input_file_path = os.path.join(directory_path, filename)
                sheet_name = os.path.splitext(filename)[0]  # Use filename (without extension) for sheet name

                # Create a new sheet
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]  # If sheet exists, use it
                else:
                    sheet = workbook.create_sheet(title=sheet_name)  # Create new sheet

                sheet.append(["Datetime", "Tput", "Gbits/sec"])  # Header row
                rx_tx_found = False  # Flag to track RX/TX occurrence
                with open(input_file_path, 'r') as infile:
                    lines = infile.readlines()
                    for line in lines:
                        if "[SUM][RX-C]" in line or "[SUM][TX-C]" in line:
                            print(f"\nSkipping {filename}: Contains RX/TX data.")
                            rx_tx_found = True
                            break  # Skip to the next file
                    if rx_tx_found:
                        pbar_files.update(1)
                        continue
                    with tqdm(total=len(lines), desc=f"Processing {filename}") as pbar:
                        for line in lines:
                            if "[SUM]" in line:
                                # Modified regex to capture bits/sec, Mbits/sec, and Gbits/sec
                                match = re.match(
                                r"(\w{3} \w{3}\s+\d{1,2} \d{2}:\d{2}:\d{2} \d{4}) \[SUM\].*?([\d\.]+) (bits|Mbits|Gbits)/sec",
                                line)
                                if match:
                                    date_str = match.group(1)
                                    transfer_str = match.group(2)
                                    unit = match.group(3)
        
                                    try:
                                        date_obj = datetime.strptime(date_str, "%a %b %d %H:%M:%S %Y")
                                        formatted_date = date_obj.strftime("%Y%m%d_%H:%M:%S")
        
                                        # Convert to Gbits/sec
                                        transfer_value = float(transfer_str)
                                        if unit == "Mbits":
                                            transfer_value /= 1000.0  # Convert Mbits to Gbits
                                        elif unit == "bits":
                                            transfer_value /= 1000000000.0 # Convert bits to Gbits
        
                                        sheet.append([formatted_date, transfer_value, 'gbps'])
                                    except ValueError:
                                        print(f"Warning: Invalid data format in line: '{line}'. Skipping.")
                            pbar.update(1)  # Update progress bar for lines
                # Adjust column width for 'Datetime' (in each sheet)
                datetime_column = sheet['A']
                max_length = 0
                for cell in datetime_column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except TypeError:
                        pass

                adjusted_width = max_length + 2
                sheet.column_dimensions['A'].width = adjusted_width
                pbar_files.update(1)  # Update progress bar for files

        workbook.save(output_excel_file)
        print(f"Data written to {output_excel_file}")

    except FileNotFoundError:
        print(f"Error: Directory '{directory_path}' not found.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

#using interval         
def plot_from_excel_simple_adjusted_interval(excel_file):
    try:
        all_sheets_data = pd.read_excel(excel_file, sheet_name=None)
        for sheet_name, df in all_sheets_data.items():
            if 'Datetime' in df.columns and 'Tput' in df.columns:
                # Ensure 'Datetime' is treated as string for direct display
                datetime_col = df['Datetime']
                tput_col = pd.to_numeric(df['Tput'], errors='coerce') # Convert Tput to numeric, handle errors
                plt.figure(figsize=(12, 5), dpi=150) # Adjust figure size if needed
                plt.plot(datetime_col, tput_col, label='Tput')
                plt.xlabel('Time')
                plt.ylabel('Throughput (gbps)')
                plt.title(f'Throughput - {sheet_name}')
                plt.legend()
                plt.grid(True)
                # Set y-axis limits
                plt.ylim(0, 4)
                # Display full datetime on x-axis
                #plt.xticks(rotation=45, ha='right')
                plt.xticks( np.linspace(0, len(datetime_col)-1, 50 ),rotation=90, ha='right' )
                plt.tight_layout()
                plt.savefig(f"tput_adjusted_{sheet_name}.png")
                plt.close()
                print(f"Adjusted plot for {sheet_name} saved.")
            else:
                print(f"Warning: Sheet '{sheet_name}' missing 'Datetime' or 'Tput' columns.")
    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

#mintute interval
def plot_from_excel_simple_adjusted(excel_file):
    try:
        all_sheets_data = pd.read_excel(excel_file, sheet_name=None)
        for sheet_name, df in all_sheets_data.items():
            if 'Datetime' in df.columns and 'Tput' in df.columns:
                # Ensure 'Datetime' is treated as string for direct display
                
                df['Datetime'] = pd.to_datetime(df['Datetime'], errors='coerce', format='%Y%m%d_%H:%M:%S')
                df.dropna(subset=['Datetime', 'Tput'], inplace=True)
                
                datetime_col = df['Datetime']
                tput_col = pd.to_numeric(df['Tput'], errors='coerce') # Convert Tput to numeric, handle errors
                
                df['Datetime'] = pd.to_datetime(df['Datetime'], errors='coerce', format='%Y%m%d_%H:%M:%S')
                df.dropna(subset=['Datetime', 'Tput'], inplace=True) # Remove rows with NaT or NaN

                datetime_col = df['Datetime']
                tput_col = pd.to_numeric(df['Tput'], errors='coerce') # Convert Tput to numeric, handle errors
                #
                '''
                plt.figure(figsize=(12, 5), dpi=150)  # Adjust figure size if needed
                plt.plot(datetime_col, tput_col, label='Tput')
                plt.xlabel('Time')
                plt.ylabel('Throughput (gbps)')
                plt.title(f'Throughput - {sheet_name}')
                plt.legend()
                plt.grid(True)
                '''
                fig, ax = plt.subplots(figsize=(12, 5), dpi=150) # Use subplots for date formatting
                ax.plot(datetime_col, tput_col, label='Tput')
                ax.set_xlabel('Time')
                ax.set_ylabel('Throughput (gbps)')
                ax.set_title(f'Throughput - {sheet_name}')
                ax.legend()
                ax.grid(True)
                
                
                '''
                # Set y-axis limits
                plt.ylim(0, 4)
                # Display full datetime on x-axis
                #plt.xticks(rotation=45, ha='right')
                
                plt.xticks( np.linspace(0, len(datetime_col)-1, 50 ),rotation=90, ha='right' )
                plt.tight_layout()
                plt.savefig(f"tput_adjusted_{sheet_name}.png")
                plt.close()
                print(f"Adjusted plot for {sheet_name} saved.")
                '''
                # Set y-axis limits
                ax.set_ylim(0, 4)
                # Set x-axis to show minute intervals
                
                ax.xaxis.set_major_locator(mdates.MinuteLocator(interval=1)) # Adjust interval as needed
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d %H:%M')) # Format as desired
                plt.xticks(rotation=45, ha='right') # Rotate labels for readability

                plt.tight_layout()
                plt.savefig(f"tput_adjusted_{sheet_name}.png")
                plt.close()
                print(f"Adjusted plot for {sheet_name} saved.")
            else:
                print(f"Warning: Sheet '{sheet_name}' missing 'Datetime' or 'Tput' columns.")
    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")
        
###############################################################
        
directory_path = "."  # Current directory (you can change this)
output_excel_file_path = "all_log_txt_output.xlsx"
print('\t\t<========================================================================>')
#process_log_files_to_excel(directory_path, output_excel_file_path)
try:
    process_log_files_to_excel(directory_path, output_excel_file_path)
except SystemExit:
    print('\nError: Your log file contains Gbits/sec or RX/TX logs. This script only captures Mbits/sec log files.')
    
print('\t\t<========================================================================>')

plot_from_excel_simple_adjusted(output_excel_file_path)
#plot_from_excel_simple_adjusted_interval(output_excel_file_path)
print("Adjusted plotting complete.")
print('\t\t<========================================================================>')