'''
- plot_from_excel_simple_adjusted(): add Dynamic Plotting Adjustments
    - 5 minute: show interval 10 second
    - 15 minute: show interval 30 second
    - 30 minute: show interval 1 minute
    - 14 hrs ~25hr: show interval 30 minute
- plot_from_excel_simple_adjusted_old(): old method all x axis just static axis time
'''
import re
import sys
from datetime import datetime, timedelta
import openpyxl
import os
from openpyxl.styles import Alignment
from tqdm import tqdm
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np

def format_duration_output(data, item_name):
    """Calculates and returns the running duration as a formatted string."""
    if not data or len(data) < 2: # Need at least two points for duration
        return f"[{item_name}] => Not enough data points to calculate duration."

    try:
        start_time_str = data[0][0]  # Assuming data is [[formatted_date, tput_value], ...]
        end_time_str = data[-1][0]

        start_time = datetime.strptime(start_time_str, "%Y%m%d_%H:%M:%S")
        end_time = datetime.strptime(end_time_str, "%Y%m%d_%H:%M:%S")

        duration = end_time - start_time
        total_seconds = int(duration.total_seconds())

        if total_seconds < 0:
            return f"[{item_name}] => Warning: End time is earlier than start time. Cannot calculate duration."

        days = total_seconds // (24 * 3600)
        remaining_seconds = total_seconds % (24 * 3600)
        hours = remaining_seconds // 3600
        remaining_seconds %= 3600
        minutes = remaining_seconds // 60
        seconds = remaining_seconds % 60

        output = [
            f"--- Duration for {item_name} ---",
            f"  Start Time: {start_time}",
            f"  End Time:   {end_time}",
            f"  Running duration: {days} days, {hours} hours, {minutes} minutes, {seconds} seconds (Total: {total_seconds}s)",
            "--------------------------------"
        ]
        return "\n".join(output)

    except ValueError:
        return f"[{item_name}] => Error parsing date/time for duration calculation. Check date format."
    except Exception as e:
        return f"[{item_name}] => An unexpected error occurred during duration calculation: {e}"


def process_log_files_to_excel(directory_path, output_excel_file):
    """
    Processes log files, extracts throughput data, writes to Excel,
    and collects duration outputs for specific data types.
    """
    all_duration_outputs = [] # List to store duration strings

    try:
        workbook = openpyxl.Workbook()

        file_list = [f for f in os.listdir(directory_path) if f.endswith((".txt", ".log"))]
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        if not file_list:
            print(f"No .txt or .log files found in directory: {directory_path}")
            return all_duration_outputs

        with tqdm(total=len(file_list), desc="Processing Files") as pbar_files:
            for filename in file_list:
                input_file_path = os.path.join(directory_path, filename)
                base_sheet_name = os.path.splitext(filename)[0]

                current_file_sheets = {
                    'default': None, # Start as None, will be created only if generic [SUM] data exists
                    'rx': None,
                    'tx': None
                }
                rx_data_for_duration = []
                tx_data_for_duration = []

                # Removed the unconditional creation of the 'default' sheet here

                log_pattern = re.compile(
                    r"(\w{3} \w{3}\s+\d{1,2} \d{2}:\d{2}:\d{2} \d{4}) " # Group 1: Datetime
                    r"\[SUM\].*?"
                    r"([\d\.]+) (bits|Mbits|Gbits)/sec"
                )

                with open(input_file_path, 'r') as infile:
                    lines = infile.readlines()
                    with tqdm(total=len(lines), desc=f"Parsing {filename}") as pbar:
                        for line in lines:
                            match = log_pattern.match(line)
                            if match:
                                date_str = match.group(1)
                                transfer_str = match.group(2)
                                unit = match.group(3)

                                try:
                                    date_obj = datetime.strptime(date_str, "%a %b %d %H:%M:%S %Y")
                                    formatted_date = date_obj.strftime("%Y%m%d_%H:%M:%S")

                                    transfer_value = float(transfer_str)
                                    if unit == "Mbits":
                                        transfer_value /= 1000.0
                                    elif unit == "bits":
                                        transfer_value /= 1000000000.0

                                    target_sheet = None # Initialize target_sheet for each line

                                    if "[SUM][RX-C]" in line:
                                        if current_file_sheets['rx'] is None:
                                            rx_sheet_name = f"{base_sheet_name}_RX"
                                            if rx_sheet_name in workbook.sheetnames:
                                                current_file_sheets['rx'] = workbook[rx_sheet_name]
                                                current_file_sheets['rx'].delete_rows(1, current_file_sheets['rx'].max_row)
                                            else:
                                                current_file_sheets['rx'] = workbook.create_sheet(title=rx_sheet_name)
                                            current_file_sheets['rx'].append(["Datetime", "Tput", "Gbits/sec"])
                                        target_sheet = current_file_sheets['rx']
                                        rx_data_for_duration.append([formatted_date, transfer_value])
                                    elif "[SUM][TX-C]" in line:
                                        if current_file_sheets['tx'] is None:
                                            tx_sheet_name = f"{base_sheet_name}_TX"
                                            if tx_sheet_name in workbook.sheetnames:
                                                current_file_sheets['tx'] = workbook[tx_sheet_name]
                                                current_file_sheets['tx'].delete_rows(1, current_file_sheets['tx'].max_row)
                                            else:
                                                current_file_sheets['tx'] = workbook.create_sheet(title=tx_sheet_name)
                                            current_file_sheets['tx'].append(["Datetime", "Tput", "Gbits/sec"])
                                        target_sheet = current_file_sheets['tx']
                                        tx_data_for_duration.append([formatted_date, transfer_value])
                                    else: # This block handles generic [SUM] lines
                                        if current_file_sheets['default'] is None: # Create default sheet ONLY if first generic SUM line is found
                                            sheet_name_default = base_sheet_name
                                            if sheet_name_default in workbook.sheetnames:
                                                current_file_sheets['default'] = workbook[sheet_name_default]
                                                current_file_sheets['default'].delete_rows(1, current_file_sheets['default'].max_row)
                                            else:
                                                current_file_sheets['default'] = workbook.create_sheet(title=sheet_name_default)
                                            current_file_sheets['default'].append(["Datetime", "Tput", "Gbits/sec"])
                                        target_sheet = current_file_sheets['default']

                                    if target_sheet: # Ensure a target_sheet was assigned before appending
                                        target_sheet.append([formatted_date, transfer_value, 'gbps'])

                                except ValueError:
                                    print(f"Warning: Invalid data format in line: '{line}'. Skipping.")
                            pbar.update(1)

                if rx_data_for_duration:
                    all_duration_outputs.append(format_duration_output(rx_data_for_duration, f"{base_sheet_name}_RX"))
                if tx_data_for_duration:
                    all_duration_outputs.append(format_duration_output(tx_data_for_duration, f"{base_sheet_name}_TX"))

                # Adjust column width for 'Datetime' in all relevant sheets for this file
                for sheet_key in current_file_sheets:
                    s = current_file_sheets[sheet_key]
                    if s: # Only adjust if the sheet was created/used
                        s.column_dimensions['A'].hidden = False
                        s.column_dimensions['A'].width = max(s.column_dimensions['A'].width, 20)

                        max_length = 0
                        for cell in s['A']:
                            try:
                                if cell.value is not None and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except TypeError:
                                pass
                        adjusted_width = max_length + 2
                        if adjusted_width > s.column_dimensions['A'].width:
                            s.column_dimensions['A'].width = adjusted_width

                pbar_files.update(1)

        if workbook.sheetnames:
            workbook.active = workbook[workbook.sheetnames[0]]

        workbook.save(output_excel_file)
        print(f"Data written to {output_excel_file}")
        return all_duration_outputs

    except FileNotFoundError:
        print(f"Error: Directory '{directory_path}' not found.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        import traceback
        traceback.print_exc()
    return all_duration_outputs

def plot_from_excel_simple_adjusted(excel_file, output_dir):
    print("\nAttempting to generate plots...")
    try:
        all_sheets_data = pd.read_excel(excel_file, sheet_name=None)
        plot_count = 0

        for sheet_name, df in all_sheets_data.items():
            print(f"Processing sheet for plotting: '{sheet_name}'")
            if 'Datetime' in df.columns and 'Tput' in df.columns:
                df['Datetime'] = pd.to_datetime(df['Datetime'], errors='coerce', format='%Y%m%d_%H:%M:%S')
                df['Tput'] = pd.to_numeric(df['Tput'], errors='coerce')
                df.dropna(subset=['Datetime', 'Tput'], inplace=True)

                if df.empty:
                    print(f"Sheet '{sheet_name}' contains no valid data after cleaning. Skipping plot generation for this sheet.")
                    continue

                datetime_col = df['Datetime']
                tput_col = df['Tput']

                # --- Dynamic Plotting Adjustments based on Data Duration ---
                min_datetime = datetime_col.min()
                max_datetime = datetime_col.max()
                total_duration = max_datetime - min_datetime

                # Default figure size and locator/interval settings
                fig_width = 12
                fig_height = 5
                major_locator = mdates.AutoDateLocator(maxticks=10) # Fallback

                # Adjust figure size and locator based on total_duration
                if total_duration <= timedelta(minutes=5): # Very short spans (e.g., few minutes)
                    fig_width = 14
                    major_locator = mdates.SecondLocator(interval=10) # Ticks every 10 seconds (up to ~30 labels)
                elif total_duration <= timedelta(minutes=15): # Your 902-second log falls here
                    fig_width = 16 # Wider plot for more labels
                    major_locator = mdates.SecondLocator(interval=30) # Ticks every 30 seconds (approx 30 labels for 15 mins)
                    # If you want even more (e.g., ~60 labels), change interval to 15, and consider making fig_width even larger
                elif total_duration <= timedelta(minutes=30): # Up to 30 minutes
                    fig_width = 18
                    major_locator = mdates.MinuteLocator(interval=1) # Ticks every 1 minute
                elif total_duration <= timedelta(hours=2): # Up to 2 hours
                    fig_width = 14
                    major_locator = mdates.MinuteLocator(interval=5) # Ticks every 5 minutes
                elif total_duration <= timedelta(hours=6): # Up to 6 hours
                    fig_width = 16
                    major_locator = mdates.MinuteLocator(interval=15) # Ticks every 15 minutes
                #Greater than 6hrs, less than or equal to 24hrs timedelta(days=1)
                elif total_duration <= timedelta(days=1): # Up to 24 hours (e.g., overnight logs like your 14hr case)
                    fig_width = 20 # Increased width to accommodate more labels
                    major_locator = mdates.MinuteLocator(interval=30) # Ticks every 30 minutes
                    # For a 14-hour log, this will give ~28 labels (14*2)
                    # For a 24-hour log, this will give ~48 labels (24*2)

 
                elif total_duration <= timedelta(days=3): # Up to 3 days
                    fig_width = 20
                    major_locator = mdates.HourLocator(interval=3) # Ticks every 3 hours
                elif total_duration <= timedelta(days=7): # Up to 7 days
                    fig_width = 22
                    major_locator = mdates.DayLocator(interval=1) # Ticks every 1 day
                else: # More than 7 days
                    fig_width = 25
                    major_locator = mdates.DayLocator(interval=7) # Ticks every 7 days (weekly)

                # Maintain a reasonable aspect ratio for height
                fig_height = fig_width * (5/12)

                fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=150)
                ax.plot(datetime_col, tput_col, label='Tput')
                ax.set_xlabel('Time')
                ax.set_ylabel('Throughput (gbps)')
                ax.set_title(f'Throughput - {sheet_name}')
                ax.legend()
                ax.grid(True)
                ax.set_ylim(0, 4.5)
                #set x axis start and end range
                ax.set_xlim(min_datetime, max_datetime) # Set x-axis limits tightly to data range
                # Apply the dynamically chosen locator and the formatter
                formatter = mdates.DateFormatter('%Y-%m-%d %H:%M:%S')
                ax.xaxis.set_major_locator(major_locator)
                ax.xaxis.set_major_formatter(formatter)
                plt.xticks(rotation=45, ha='right')

                plt.tight_layout()
                png_output_path = os.path.join(output_dir, f"tput_adjusted_{sheet_name}.png")
                plt.savefig(png_output_path)
                plt.close(fig)
                print(f"Plot for '{sheet_name}' saved as '{png_output_path}'")
                plot_count += 1
            else:
                print(f"Warning: Sheet '{sheet_name}' missing 'Datetime' or 'Tput' columns. Skipping plot.")

        if plot_count == 0:
            print("No plots were generated for any sheet due to lack of valid data.")

    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file}' not found for plotting.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        import traceback
        traceback.print_exc()
        
# --- Plotting Function (No changes needed here for this feature) ---
def plot_from_excel_simple_adjusted_old(excel_file, output_dir):
    print("\nAttempting to generate plots...")
    try:
        all_sheets_data = pd.read_excel(excel_file, sheet_name=None)
        plot_count = 0

        for sheet_name, df in all_sheets_data.items():
            print(f"Processing sheet for plotting: '{sheet_name}'")
            if 'Datetime' in df.columns and 'Tput' in df.columns:
                df['Datetime'] = pd.to_datetime(df['Datetime'], errors='coerce', format='%Y%m%d_%H:%M:%S')
                df['Tput'] = pd.to_numeric(df['Tput'], errors='coerce')
                df.dropna(subset=['Datetime', 'Tput'], inplace=True)

                if df.empty:
                    print(f"Sheet '{sheet_name}' contains no valid data after cleaning. Skipping plot generation for this sheet.")
                    continue

                datetime_col = df['Datetime']
                tput_col = df['Tput']

                fig, ax = plt.subplots(figsize=(12, 5), dpi=150)
                ax.plot(datetime_col, tput_col, label='Tput')
                ax.set_xlabel('Time')
                ax.set_ylabel('Throughput (gbps)')
                ax.set_title(f'Throughput - {sheet_name}')
                ax.legend()
                ax.grid(True)
                ax.set_ylim(0, 4.5) #y axis range

                locator = mdates.AutoDateLocator(maxticks=200)
                formatter = mdates.DateFormatter('%Y-%m-%d %H:%M:%S')
                ax.xaxis.set_major_locator(locator)
                ax.xaxis.set_major_formatter(formatter)
                plt.xticks(rotation=45, ha='right')

                plt.tight_layout()
                png_output_path = os.path.join(output_dir, f"tput_adjusted_{sheet_name}.png")
                plt.savefig(png_output_path)
                plt.close(fig)
                print(f"Plot for '{sheet_name}' saved as '{png_output_path}'")
                plot_count += 1
            else:
                print(f"Warning: Sheet '{sheet_name}' missing 'Datetime' or 'Tput' columns. Skipping plot.")

        if plot_count == 0:
            print("No plots were generated for any sheet due to lack of valid data.")

    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file}' not found for plotting.")
    except Exception as e:
        print(f"An unexpected error occurred during plotting: {e}")
        import traceback
        traceback.print_exc()

# --- Main Execution Block ---
if __name__ == "__main__":
    directory_path = "."
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    results_folder_name = f"{timestamp}_results"
    results_folder_path = os.path.join(directory_path, results_folder_name)

    try:
        os.makedirs(results_folder_path, exist_ok=True)
        print(f"Created results directory: {results_folder_path}")
    except OSError as e:
        print(f"Error creating results directory {results_folder_path}: {e}")
        sys.exit(1)

    output_excel_file_path = os.path.join(results_folder_path, f"all_logs_analysis_{timestamp}.xlsx")

    print(f"\t\t<==================== Starting Log Processing ====================>")
    print(f"Output Excel will be saved to: {output_excel_file_path}")
    print(f"Processing logs from directory: {os.path.abspath(directory_path)}")

    collected_duration_outputs = process_log_files_to_excel(directory_path, output_excel_file_path)

    print(f"\t\t<==================== Log Processing Finished ====================>")

    if os.path.exists(output_excel_file_path):
        plot_from_excel_simple_adjusted(output_excel_file_path, results_folder_path)
        print("\nPlotting process complete.")
    else:
        print(f"\nExcel file '{output_excel_file_path}' was not created or found. Skipping plotting.")

    if collected_duration_outputs:
        print("\n\n<==================== Duration Summaries ====================>")
        for duration_str in collected_duration_outputs:
            print(duration_str)
            print()
        print("<===========================================================>")
    else:
        print("\nNo specific RX/TX durations to display.")


    print("\nScript finished.")